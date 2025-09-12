from __future__ import annotations
import re
import time
import os
import shutil
from typing import Iterable, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def read_allocation_pricesheet(folder: str = "put_your_excel_here"):
    """
    Returns
    -------
    allocation_df : pd.DataFrame | None
        The allocation workbook's default/active VISIBLE sheet (what opens first in Excel),
        raw (header=None). Returns None if not found.
    price_sheet_df : pd.DataFrame | None
        The price workbook's 'script' sheet, raw (header=None). Returns None if not found.

    Notes
    -----
    - Accepts one or both files:
        * One file whose name contains 'allocation'
        * One file whose name contains 'price'
    - Temp/lock files like '~$*.xlsx' are ignored.
    """
    folder_path = Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path.resolve()}")

    excel_paths = _find_excel_files(folder_path)
    if len(excel_paths) == 0 or len(excel_paths) > 2:
        raise ValueError(
            f"Expected 1–2 Excel files, found {len(excel_paths)} in {folder_path.resolve()}.\n"
            f"Files seen: {[p.name for p in excel_paths]}"
        )

    alloc_path = _pick_file_by_keyword(excel_paths, "allocation")
    price_path = _pick_file_by_keyword(excel_paths, "price")

    allocation_df = None
    price_sheet_df = None

    if alloc_path is not None:
        # Allocation: read the workbook's default (active) visible sheet
        alloc_sheet = _get_active_visible_sheet_name(alloc_path)
        allocation_df = pd.read_excel(
            alloc_path, sheet_name=alloc_sheet, header=None, engine="openpyxl"
        )

    if price_path is not None:
        # Price: read 'script'
        _assert_visible_sheet(price_path, "script")
        price_sheet_df = pd.read_excel(
            price_path, sheet_name="script", header=None, engine="openpyxl"
        )

    if allocation_df is None and price_sheet_df is None:
        raise ValueError(
            "Neither 'allocation' nor 'price' file found (case-insensitive)."
        )

    return allocation_df, price_sheet_df


# -----------------------
# Helpers
# -----------------------

def _find_excel_files(folder_path: Path):
    """*.xlsx/*.xlsm/*.xls, excluding Office temp/lock files like '~$…'."""
    pats = ("*.xlsx", "*.xlsm", "*.xls")
    files = []
    for pat in pats:
        files.extend(folder_path.glob(pat))
    return [p for p in files if p.is_file() and not p.name.startswith("~$")]

def _pick_file_by_keyword(paths, keyword: str):
    """First file whose stem contains keyword (case-insensitive)."""
    kw = keyword.lower()
    for p in paths:
        if kw in p.stem.lower():
            return p
    return None

def _visible_sheet_names(xlsx_path: Path):
    """List titles of visible sheets only (ignore hidden/veryHidden)."""
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    finally:
        wb.close()

def _get_active_visible_sheet_name(xlsx_path: Path) -> str:
    """
    Return the active sheet's name if it's visible; otherwise fall back to the
    first visible sheet. Raise if none are visible.
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        active = wb.active
        if getattr(active, "sheet_state", "visible") == "visible":
            return active.title
        # Fallback: first visible sheet
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") == "visible":
                return ws.title
        raise ValueError(f"No visible sheets in {xlsx_path.name}.")
    finally:
        wb.close()

def _assert_visible_sheet(xlsx_path: Path, sheet_name: str):
    vis = _visible_sheet_names(xlsx_path)
    if sheet_name not in vis:
        raise ValueError(
            f"Sheet '{sheet_name}' not found as a VISIBLE tab in {xlsx_path.name}.\n"
            f"Visible sheets: {vis}"
        )

def read_latest_po_csv(
    folder: str = r"C:\POs",
    delete_after: bool = False,
) -> pd.DataFrame:
    """
    Find the most recently modified .csv in `folder` (ignoring '~$' and hidden files),
    read it as one PO per line (no header, no delimiter sniffing), print PO #s, and
    return a cleaned DataFrame with columns:
        - 'PO #': original string (trimmed)
        - 'Store': left of first dash
        - 'Item' : right of first dash
    Rows are dropped if empty/NA-like OR without a dash (supports -, – or —).
    """
    folder_path = Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    # Collect candidate CSVs
    csv_files = [
        p for p in folder_path.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".csv"
        and not p.name.startswith(("~$", "."))
    ]
    if not csv_files:
        raise FileNotFoundError(f"No .csv files found in {folder_path}")

    # Pick most recently modified
    file_path = max(csv_files, key=lambda p: p.stat().st_mtime)
    mtime = datetime.fromtimestamp(file_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    print(f"Reading latest file: {file_path.name} (modified {mtime})")

    # --- Robust single-column read: treat file as "one PO per line" ---
    lines = None
    for enc in ("utf-8-sig", "utf-16", "latin1"):
        try:
            with open(file_path, "r", encoding=enc, errors="strict") as f:
                lines = [ln.strip() for ln in f.read().splitlines()]
            break
        except UnicodeError:
            continue
    if lines is None:
        # Last resort: ignore decoding errors
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [ln.strip() for ln in f.read().splitlines()]

    raw = pd.DataFrame(lines, columns=[0])

    # First column only -> Series of strings
    s = raw.iloc[:, 0].astype(str).str.strip()

    # Clean/filter
    NA_STRINGS = {"", "na", "n/a", "nan", "none", "null", "nah"}
    has_value = ~s.str.lower().isin(NA_STRINGS)
    has_dash = s.str.contains(r"[-–—]", regex=True, na=False)
    s = s[has_value & has_dash]

    # Split into Store / Item on first dash-like char
    parts = s.str.split(r"[-–—]", n=1, expand=True)
    po_df = pd.DataFrame({
        "PO #": s.values,
        "Store": parts[0].str.strip(),
        "Item":  parts[1].str.strip()
    }).reset_index(drop=True)

    # Print POs
    if not po_df.empty:
        print("Received POs:")
        for po in po_df["PO #"]:
            print(po)
    else:
        print("No valid PO rows found (after cleaning).")

    return po_df


def retrieve_pdf(
    po_df: pd.DataFrame,
    output_folder: str | Path,
    *watch_folders: str | Path,
    pdf_folder: Optional[str | Path] = "pdf_folder",   # can be simple name OR full path
    po_col_candidates: Iterable[str] = ("PO #", "PO#", "PO_Number", "PO"),
    poll_interval: float = 5.0,
    settle_time: float = 3.0,
    open_retry: int = 5,
    open_retry_sleep: float = 1.0,
    max_wait_seconds: Optional[int] = None,
    case_insensitive: bool = True,
    verbose: bool = True,
) -> pd.DataFrame:
    """
    Monitor watch_folders for PDFs whose names end with `-<PO>.pdf` and move them to the destination dir.

    Destination resolution:
      - If `pdf_folder` is None: destination = <output_folder>
      - If `pdf_folder` is a simple name (no path separators, not absolute):
          destination = <output_folder>/<pdf_folder>
      - If `pdf_folder` is a full/absolute path OR contains path separators:
          destination = <pdf_folder>  (used as-is; NOT joined with output_folder)

    The destination directory is created only if it does not already exist.
    Also:
      - Pre-check destination to mark already-present PO PDFs as done.
      - Initial sweep of watch folders before entering the polling loop.
    """
    # ---- 1) Pick PO column
    po_col = next((c for c in po_col_candidates if c in po_df.columns), None)
    if po_col is None:
        raise ValueError(f"po_df must contain one of columns: {po_col_candidates}")

    # ---- 2) Normalize POs
    raw_pos = (
        po_df[po_col]
        .astype(str)
        .str.strip()
        .replace({"": None})
        .dropna()
        .tolist()
    )
    target_pos = list(dict.fromkeys(raw_pos))
    if not target_pos:
        raise ValueError("No valid PO values found in po_df.")

    # ---- 3) Resolve destination directory (NO extra nesting)
    output_folder = Path(output_folder)

    def is_simple_name(p: str | Path) -> bool:
        p = str(p)
        return (os.sep not in p) and (os.altsep not in p if os.altsep else True) and (not os.path.isabs(p))

    if pdf_folder is None:
        dest_dir = output_folder
    else:
        pdf_folder = Path(pdf_folder)
        if pdf_folder.is_absolute() or not is_simple_name(pdf_folder):
            dest_dir = Path(pdf_folder)           # use as-is
        else:
            dest_dir = output_folder / pdf_folder # join to output_folder

    # Create ONLY the destination dir (and its parents) if missing
    if not dest_dir.exists():
        dest_dir.mkdir(parents=True, exist_ok=True)

    # ---- 4) Watch dirs
    watch_dirs = [Path(p) for p in watch_folders]
    if not watch_dirs:
        raise ValueError("Please provide at least one watch folder.")
    for d in watch_dirs:
        if d.exists() and not d.is_dir():
            raise ValueError(f"Not a directory: {d}")

    # ---- 5) Compile patterns: '-<PO>.pdf' at filename end
    def compile_pattern(po: str):
        esc = re.escape(po)
        flags = re.IGNORECASE if case_insensitive else 0
        return re.compile(rf"-{esc}\.pdf$", flags)

    patterns = {po: compile_pattern(po) for po in target_pos}

    # ---- 6) Status DF
    status = pd.DataFrame(
        {
            "PO": target_pos,
            "found_path": [None] * len(target_pos),
            "moved_to": [None] * len(target_pos),
            "status": ["waiting"] * len(target_pos),
            "finished_at": [None] * len(target_pos),
        }
    )

    def idx_of(po: str) -> int:
        return int(status.index[status["PO"] == po][0])

    # ---- 7) Stable-file check
    def is_file_stable(p: Path, window: float) -> bool:
        try:
            size1 = p.stat().st_size
        except FileNotFoundError:
            return False
        t0 = time.time()
        while time.time() - t0 < window:
            time.sleep(0.5)
            try:
                size2 = p.stat().st_size
            except FileNotFoundError:
                return False
            if size2 != size1:
                size1 = size2
                t0 = time.time()
        return True

    # ---- 8) PRE-CHECK: mark POs already present in destination as done
    try:
        for existing in dest_dir.iterdir():
            if not existing.is_file() or not existing.name.lower().endswith(".pdf"):
                continue
            for po in target_pos:
                if status.loc[status["PO"] == po, "status"].item() == "done":
                    continue
                if patterns[po].search(existing.name):
                    i = idx_of(po)
                    status.at[i, "found_path"] = str(existing)
                    status.at[i, "moved_to"] = str(existing)
                    status.at[i, "status"] = "done"
                    status.at[i, "finished_at"] = datetime.now().isoformat(timespec="seconds")
                    break
    except Exception as e:
        if verbose:
            print(f"[WARN] Destination pre-check failed: {e}")

    remaining = set(status.loc[status["status"] != "done", "PO"].tolist())

    # Helper for logging
    def print_progress():
        if not verbose:
            return
        waiting = sorted(remaining)
        done = status[status["status"] == "done"]["PO"].tolist()
        print(
            f"[{datetime.now().strftime('%H:%M:%S')}] "
            f"waiting: {len(waiting)} -> {waiting[:10]}{'...' if len(waiting) > 10 else ''} | "
            f"done: {len(done)}"
        )

    print_progress()

    # ---- 9) INITIAL SWEEP: move any already-existing matches in watch folders
    def sweep_once():
        nonlocal remaining
        any_found = False
        for wd in watch_dirs:
            if not wd.exists() or not wd.is_dir():
                continue
            try:
                for entry in wd.iterdir():
                    if not entry.is_file() or not entry.name.lower().endswith(".pdf"):
                        continue
                    # If this file is already in dest_dir (same path), skip
                    if entry.parent.resolve() == dest_dir.resolve():
                        continue

                    for po in list(remaining):
                        if patterns[po].search(entry.name):
                            any_found = True
                            i = idx_of(po)
                            status.at[i, "found_path"] = str(entry)

                            if verbose:
                                print(f"[HIT] (initial sweep) PO={po} at {entry}")
                                print(f"      Waiting for stability: {settle_time}s")
                            if not is_file_stable(entry, settle_time):
                                if verbose:
                                    print(f"[WARN] Unstable or vanished: {entry}")
                                continue

                            # If a same-named file already exists in destination, don't move; just mark done.
                            dest = dest_dir / entry.name
                            if dest.exists():
                                status.at[i, "moved_to"] = str(dest)
                                status.at[i, "status"] = "done"
                                status.at[i, "finished_at"] = datetime.now().isoformat(timespec="seconds")
                                remaining.discard(po)
                                if verbose:
                                    print(f"[SKIP MOVE] Already in destination: {dest} — marked done")
                                break

                            # Try to move
                            ok = False
                            last_err = None
                            for _ in range(open_retry):
                                try:
                                    shutil.move(str(entry), str(dest))
                                    ok = True
                                    break
                                except Exception as e:
                                    last_err = e
                                    time.sleep(open_retry_sleep)

                            if not ok:
                                if verbose:
                                    print(f"[ERROR] Move failed: {entry} -> {dest} | {last_err}")
                                continue

                            status.at[i, "moved_to"] = str(dest)
                            status.at[i, "status"] = "done"
                            status.at[i, "finished_at"] = datetime.now().isoformat(timespec="seconds")
                            remaining.discard(po)

                            if verbose:
                                print(f"[OK] Moved (initial sweep): {entry.name} -> {dest}")
                            break
            except PermissionError:
                if verbose:
                    print(f"[WARN] Permission denied (temporary): {wd}")
            except FileNotFoundError:
                if verbose:
                    print(f"[WARN] Directory unavailable (temporary): {wd}")
            except Exception as e:
                if verbose:
                    print(f"[WARN] Initial sweep error: {wd} | {e}")
        return any_found

    initial_found = sweep_once()
    if initial_found and verbose:
        print_progress()

    # If everything is already satisfied, return early
    if not remaining:
        if verbose:
            print(f"[DONE] All POs satisfied from destination and initial sweep. Output dir: {dest_dir}")
        return status

    # ---- 10) POLLING LOOP
    start = time.time()
    while remaining:
        if max_wait_seconds is not None and (time.time() - start) > max_wait_seconds:
            if verbose:
                print("[INFO] Reached max_wait_seconds. Returning current status.")
            break

        any_found = False

        for wd in watch_dirs:
            if not wd.exists() or not wd.is_dir():
                continue

            try:
                for entry in wd.iterdir():
                    if not entry.is_file() or not entry.name.lower().endswith(".pdf"):
                        continue
                    if entry.parent.resolve() == dest_dir.resolve():
                        continue

                    for po in list(remaining):
                        if patterns[po].search(entry.name):
                            any_found = True
                            i = idx_of(po)
                            status.at[i, "found_path"] = str(entry)

                            if verbose:
                                print(f"[HIT] PO={po} at {entry}")
                                print(f"      Waiting for stability: {settle_time}s")
                            if not is_file_stable(entry, settle_time):
                                if verbose:
                                    print(f"[WARN] Unstable or vanished: {entry}")
                                continue

                            dest = dest_dir / entry.name
                            if dest.exists():
                                status.at[i, "moved_to"] = str(dest)
                                status.at[i, "status"] = "done"
                                status.at[i, "finished_at"] = datetime.now().isoformat(timespec="seconds")
                                remaining.discard(po)
                                if verbose:
                                    print(f"[SKIP MOVE] Already in destination: {dest} — marked done")
                                break

                            ok = False
                            last_err = None
                            for _ in range(open_retry):
                                try:
                                    shutil.move(str(entry), str(dest))
                                    ok = True
                                    break
                                except Exception as e:
                                    last_err = e
                                    time.sleep(open_retry_sleep)

                            if not ok:
                                if verbose:
                                    print(f"[ERROR] Move failed: {entry} -> {dest} | {last_err}")
                                continue

                            status.at[i, "moved_to"] = str(dest)
                            status.at[i, "status"] = "done"
                            status.at[i, "finished_at"] = datetime.now().isoformat(timespec="seconds")
                            remaining.discard(po)

                            if verbose:
                                print(f"[OK] Moved: {entry.name} -> {dest}")
                            break

            except PermissionError:
                if verbose:
                    print(f"[WARN] Permission denied (temporary): {wd}")
            except FileNotFoundError:
                if verbose:
                    print(f"[WARN] Directory unavailable (temporary): {wd}")
            except Exception as e:
                if verbose:
                    print(f"[WARN] Scan error: {wd} | {e}")

        if remaining:
            if any_found:
                print_progress()
            time.sleep(poll_interval)

    if verbose:
        done = status[status["status"] == "done"].shape[0]
        total = status.shape[0]
        print(f"[DONE] Completed {done}/{total}. Output dir: {dest_dir}")

    return status
