import pandas as pd
import numpy as np
import re
from pathlib import Path
from datetime import date, timedelta
from openpyxl.styles import numbers
from openpyxl import load_workbook

# ---------- helpers ----------
def _norm(s: pd.Series) -> pd.Series:
    return (s.astype(str).str.strip().str.lower()
            .str.replace(r"\s+", "", regex=True).str.replace("#", "", regex=False))

def _norm_cell(x):
    return str(x).strip().lower().replace(" ", "")

def _leading_num(x):
    if pd.isna(x): return 0
    s = str(x)
    m = re.match(r'^\s*\$?\s*([+-]?\d[\d,]*\.?\d*)', s)
    if not m: return 0
    try: return float(m.group(1).replace(',', ''))
    except ValueError: return 0

def _first_int_in_text(s):
    m = re.search(r'\d+', str(s))
    return int(m.group()) if m else None

def _norm_name(s):
    return str(s).strip().lower().replace(" ", "")

def _num_anywhere(x):
    if pd.isna(x): return 0.0
    s = str(x)
    m = re.search(r'[+-]?\d[\d,]*\.?\d*', s)
    if not m: return 0.0
    return float(m.group().replace(',', '')) if m.group() not in {'', '-', '.', '-.', '.-'} else 0.0

def _first_int_or_nan(s):
    m = re.search(r'\d+', str(s))
    return int(m.group()) if m else np.nan

def _norm_header(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def _norm_key(s):
    return str(s).strip().lower().replace(" ", "")

def _first_int_or_zero(s):
    m = re.search(r'\d+', str(s))
    return int(m.group()) if m else 0
    

# ---------- split ----------
def split_big_and_baby(df: pd.DataFrame):
    """
    - big_flip_df: unchanged logic -> from top through the first row where col 4 == 'Total Weight'
    - baby_flip_df: NEW logic -> from the **second** 'Item' (in col 1) to the end of the sheet
    """
    if df.shape[1] < 4:
        raise ValueError("Expected at least 4 columns in df.")

    # Normalize only for boundary detection (does NOT mutate df)
    col1 = _norm(df.iloc[:, 0])
    col4 = _norm(df.iloc[:, 3])

    # ---- big part boundary (unchanged) ----
    tw_pos_arr = np.where(col4.eq("totalweight").to_numpy())[0]
    if tw_pos_arr.size == 0:
        raise ValueError("No row where col4 == 'Total Weight'.")
    tw_pos = int(tw_pos_arr[0])
    big_flip_df = df.iloc[:tw_pos].copy()

    # ---- baby part boundary (revised) ----
    item_pos_arr = np.where(col1.eq("item").to_numpy())[0]
    if item_pos_arr.size < 2:
        raise ValueError(f"Need at least two 'Item' markers in first column; found {item_pos_arr.size}.")
    item2_pos = int(item_pos_arr[1])  # second occurrence
    baby_flip_df = df.iloc[item2_pos:].copy()  # from 2nd 'Item' through end

    return big_flip_df, baby_flip_df

# ---------- store ----------
def build_big_flip_store(
    big_flip_df: pd.DataFrame,
    start_col: int = 4,
    lot_label: str = "Lot #",
    fallback_left_label: str = "Total",
    lot_row_idx: int = 4,
    header_row_idx: int = 4
) -> pd.DataFrame:
    cdf = big_flip_df.copy(deep=True)
    lot_norm = _norm_cell(lot_label)
    total_norm = _norm_cell(fallback_left_label)
    stop_col = None
    for c in range(start_col, cdf.shape[1]):
        if _norm_cell(cdf.iat[lot_row_idx, c]) == lot_norm:
            stop_col = c; break
    if stop_col is None:
        for c in range(start_col, cdf.shape[1]):
            if _norm_cell(cdf.iat[lot_row_idx, c]) == total_norm:
                stop_col = c + 1; break
    if stop_col is None or stop_col <= start_col:
        raise ValueError("Neither 'Lot #' nor 'Total' found on row 5 at/after column E.")
    store = cdf.iloc[:, start_col:stop_col].copy()
    raw_header = store.iloc[header_row_idx, :]
    valid_pos = [i for i, v in enumerate(raw_header) if (pd.notna(v) and str(v).strip() != "")]
    start_row = max(0, header_row_idx - 4)
    end_row = header_row_idx
    store = store.iloc[start_row:end_row, valid_pos].copy()
    store.columns = [str(raw_header.iloc[i]).strip() for i in valid_pos]
    cols_norm = {c: _norm_cell(c) for c in store.columns}
    total_cols = [c for c, n in cols_norm.items() if n == total_norm]
    if total_cols:
        store = store.drop(columns=total_cols)
    new_cols = []
    for c in store.columns:
        n = _first_int_in_text(c)
        new_cols.append(n if n is not None else c)
    store.columns = new_cols
    store = store.reset_index(drop=True)
    to_drop = [i for i in (1, 3) if i < len(store)]
    store = store.drop(index=to_drop).reset_index(drop=True)
    store.insert(0, "Name", "")
    if len(store) > 0: store.at[0, "Name"] = "Fob"
    if len(store) > 1: store.at[1, "Name"] = "Xdock"
    value_cols = [c for c in store.columns if c != "Name"]
    store[value_cols] = store[value_cols].applymap(_leading_num)
    return store

# ---------- CLEAN: keep 'Lot #' column ----------
def clean_big_flip_df(
    big_flip_df: pd.DataFrame,
    po_header: str = "PO #",
    lot_header: str = "Lot #",
    total_header: str = "Total",
) -> pd.DataFrame:
    df = big_flip_df.copy(deep=True)
    # 1) drop rows 0..3
    if len(df) >= 1:
        df = df.drop(index=[i for i in range(min(4, len(df)))], errors="ignore").reset_index(drop=True)
    #    drop columns 1..3 (keep col0 and >=4)
    if df.shape[1] > 0:
        keep_pos = [0] + list(range(4, df.shape[1]))
        keep_pos = [p for p in keep_pos if p < df.shape[1]]
        df = df.iloc[:, keep_pos]
    # 2) header
    if df.shape[0] == 0: return df
    raw_header = df.iloc[0, :]
    # 3) drop empty header cols
    valid_pos = [i for i, v in enumerate(raw_header) if (pd.notna(v) and str(v).strip() != "")]
    if not valid_pos: return pd.DataFrame()
    df = df.iloc[1:, valid_pos].copy()
    df.columns = [str(raw_header.iloc[i]).strip() for i in valid_pos]
    # 5) drop rows where first col empty
    if df.shape[1] == 0: return df
    first_col = df.columns[0]
    mask = df[first_col].notna() & (df[first_col].astype(str).str.strip() != "")
    df = df[mask].copy()
    # 6) right-trim: PO# -> left of; Lot# -> INCLUDE Lot#; Total -> include
    norms = [_norm_header(c) for c in df.columns]
    po_norm   = _norm_header(po_header)
    lot_norm  = _norm_header(lot_header)
    total_norm= _norm_header(total_header)
    def _cut_left_of(norm_key):
        idx = norms.index(norm_key)
        return df.iloc[:, :idx].copy() if idx > 0 else pd.DataFrame()
    if po_norm in norms:
        df = _cut_left_of(po_norm)
    elif lot_norm in norms:
        idx = norms.index(lot_norm)
        df = df.iloc[:, : idx + 1].copy()    # <-- keep 'Lot #'
    elif total_norm in norms:
        idx = norms.index(total_norm)
        df = df.iloc[:, : idx + 1].copy()    # include 'Total'
    df = df.reset_index(drop=True)
    return df

# ---------- PIVOT: group by Branch, Item, Lot # ----------
def build_big_flip_df_cleaned_pivot(big_flip_df_cleaned: pd.DataFrame, item_col_name: str | None = None) -> pd.DataFrame:
    # identify Item
    if item_col_name is None:
        norms = [_norm_name(c) for c in big_flip_df_cleaned.columns]
        if 'item' in norms:
            item_col_name = big_flip_df_cleaned.columns[norms.index('item')]
        else:
            raise ValueError("Couldn't find an 'Item' column. Pass item_col_name='YourItemHeader'.")
    # require 'Lot #'
    lot_candidates = [c for c in big_flip_df_cleaned.columns if _norm_header(c) == _norm_header("Lot #")]
    if not lot_candidates:
        raise ValueError("'Lot #' column not found in big_flip_df_cleaned. Ensure clean_big_flip_df kept it.")
    lot_col = lot_candidates[0]

    # melt branch columns (exclude Item & Lot #)
    value_cols = [c for c in big_flip_df_cleaned.columns if c not in (item_col_name, lot_col)]
    long = big_flip_df_cleaned.melt(
        id_vars=[item_col_name, lot_col],
        value_vars=value_cols,
        var_name='Branch',
        value_name='raw_value'
    )

    # numeric, aggregate by Branch, Item, Lot #
    long['Distro Size'] = long['raw_value'].apply(_num_anywhere)
    out = long.groupby(['Branch', item_col_name, lot_col], as_index=False)['Distro Size'].sum()
    out['Distro Size'] = np.ceil(out['Distro Size'].fillna(0)).astype(int)
    out = out[out['Distro Size'] != 0].copy()

    # sort
    out['__branch_num'] = out['Branch'].apply(_first_int_or_nan)
    out = (out.sort_values(
            by=['__branch_num', 'Branch', item_col_name, lot_col, 'Distro Size'],
            ascending=[True, True, True, True, True],
            na_position='last')
           .drop(columns='__branch_num')
           .reset_index(drop=True))

    # final columns
    out = out.rename(columns={item_col_name: 'Item'})
    out = out[['Branch', 'Item', lot_col, 'Distro Size']]  # keep 'Lot #'
    return out

# ---------- LOOKUPS & OUTPUT (unchanged schema) ----------
def _make_lookup_maps(big_flip_store: pd.DataFrame):
    if 'Name' not in big_flip_store.columns:
        raise ValueError("big_flip_store must have a 'Name' column (values 'Fob' and 'Xdock').")
    fob_row  = big_flip_store[big_flip_store['Name'].astype(str).str.strip().str.lower() == 'fob']
    xdk_row  = big_flip_store[big_flip_store['Name'].astype(str).str.strip().str.lower() == 'xdock']
    if fob_row.empty or xdk_row.empty: return {}, {}
    fob_row = fob_row.iloc[0]; xdk_row = xdk_row.iloc[0]
    branch_cols = [c for c in big_flip_store.columns if c != 'Name']
    fob_map, xdock_map = {}, {}
    for c in branch_cols:
        fob_map[c] = fob_row[c]; xdock_map[c] = xdk_row[c]
        fob_map[_norm_key(c)] = fob_row[c]; xdock_map[_norm_key(c)] = xdk_row[c]
    return xdock_map, fob_map

def _safe_lookup(map_obj, key):
    if key in map_obj: v = map_obj[key]
    else:              v = map_obj.get(_norm_key(key), np.nan)
    try:
        if pd.isna(v) or (isinstance(v, str) and v.strip() == ""): return ""
        v_float = float(v); return "" if v_float == 0 else v_float
    except Exception:
        return ""
        
def _next_mwf_date_str(tz="America/Chicago"):
    # Next shipping day among Mon(0), Wed(2), Fri(4)
    today = pd.Timestamp.now(tz=tz).normalize()
    dow = today.dayofweek  # Mon=0 ... Sun=6
    if dow == 0:      delta = 2   # Mon -> Wed
    elif dow == 2:    delta = 2   # Wed -> Fri
    elif dow == 4:    delta = 3   # Fri -> next Mon
    elif dow == 1:    delta = 1   # Tue -> Wed
    elif dow == 3:    delta = 1   # Thu -> Fri
    elif dow == 5:    delta = 2   # Sat -> Mon
    else:             delta = 1   # Sun -> Mon
    target = today + pd.Timedelta(days=delta)
    return target.strftime("%m/%d/%y")
    
def build_big_flip_output(
    big_flip_df_cleaned_pivot: pd.DataFrame,
    big_flip_store: pd.DataFrame,
    tz: str = "America/Chicago"
) -> pd.DataFrame:
    required_any = {'Branch', 'Item', 'Distro Size'}
    miss = required_any - set(big_flip_df_cleaned_pivot.columns)
    if miss:
        raise ValueError(f"big_flip_df_cleaned_pivot missing columns: {miss}")
    out = big_flip_df_cleaned_pivot[['Branch','Item','Distro Size']].copy()
    out['Branch'] = out['Branch'].apply(_first_int_or_zero).astype(int)
    out['Item']   = out['Item'].apply(_first_int_or_zero).astype(int)
    out['Description'] = ""
    out['Supplier On Record'] = ""
    out['Warehouse'] = ""
    out['AdditionalXDCK'] = ""
    out['POSTXDCK'] = ""
    out['WW Buyer'] = "P20"
    out['AmountCode'] = "W"
    out['Expected Delivery Date'] = _next_mwf_date_str(tz=tz)
    xdock_map, fob_map = _make_lookup_maps(big_flip_store)
    out['XDCK'] = out['Branch'].apply(lambda b: _safe_lookup(xdock_map, b))
    out['FOB']  = out['Branch'].apply(lambda b: _safe_lookup(fob_map,  b))
    str_cols = ['Description','Supplier On Record','Expected Delivery Date','WW Buyer',
                'Warehouse','AdditionalXDCK','AmountCode','XDCK','POSTXDCK','FOB']
    out[str_cols] = out[str_cols].fillna("")
    out['Distro Size'] = pd.to_numeric(out['Distro Size'], errors='coerce').fillna(0).astype(int)
    cols = ['Branch','Item','Description','Distro Size','Supplier On Record','Expected Delivery Date',
            'WW Buyer','Warehouse','AdditionalXDCK','AmountCode','XDCK','POSTXDCK','FOB']
    out = out[cols]
    out = out.sort_values(by=['Branch','Item','Distro Size'], ascending=[True, True, True]).reset_index(drop=True)
    return out

CANONICAL_COLS = [
    'Branch','Item','Description','Distro Size','Supplier On Record','Expected Delivery Date',
    'WW Buyer','Warehouse','AdditionalXDCK','AmountCode','XDCK','POSTXDCK','FOB'
]

def write_big_flip_output_excel(df: pd.DataFrame, path: str) -> None:
    df = df.copy().reindex(columns=CANONICAL_COLS)

    # numeric cols
    for c in ['Branch','Item','Distro Size']:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).astype(int)
    for c in ['XDCK','FOB']:
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # ensure real dates (accepts both 2- and 4-digit year strings)
    if 'Expected Delivery Date' in df.columns:
        df['Expected Delivery Date'] = pd.to_datetime(df['Expected Delivery Date'], errors='coerce').dt.date

    # text -> blanks
    for c in ['Description','Supplier On Record','WW Buyer','Warehouse','AdditionalXDCK','AmountCode','POSTXDCK']:
        df[c] = df[c].astype(object).where(df[c].notna(), "")

    # sort
    df = df.sort_values(by=['Branch','Item','Distro Size'], ascending=[True, True, True]).reset_index(drop=True)

    # write workbook/sheets
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Scripting')
        writer.book.create_sheet("ANOMALY")
        writer.book.create_sheet("STORE CLUSTER")

        # >>> Format the date column as m/d/yyyy (no leading zeros, 4-digit year)
        ws = writer.book['Scripting']
        date_col_idx = df.columns.get_loc('Expected Delivery Date') + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=date_col_idx)
            if cell.value is not None:
                cell.number_format = "m/d/yyyy"

        # blank NaNs in XDCK/FOB
        for name in ['XDCK','FOB']:
            ci = df.columns.get_loc(name) + 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=ci).value is None:
                    ws.cell(row=r, column=ci).value = ""


def build_big_flip_output_path(file_name: str, folder: str = "output_folder") -> Path:
    """
    Creates: <folder>/<file_name> IBTs <MM-DD-YY>.xlsx
    - trims/collapses spaces in file_name
    - ensures .xlsx suffix
    """
    today_str = date.today().strftime("%m-%d-%y")  # e.g., 08-29-25
    base = re.sub(r"\s+", " ", str(file_name)).strip()
    path = Path(folder) / f"{base} Big Flip {today_str}.xlsx"
    return path